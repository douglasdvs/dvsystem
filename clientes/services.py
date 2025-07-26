import logging
from datetime import datetime
from functools import lru_cache
from typing import Dict, List, Optional

import requests
from django.conf import settings
from django.core.cache import cache
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Cliente

logger = logging.getLogger(__name__)


@lru_cache(maxsize=1)
def ufs_choices():
    """Retorna lista de UFs disponíveis (cacheada)"""
    return [uf for uf, _ in Cliente._meta.get_field("uf").choices]


def buscar_clientes_por_usuario(user):
    """Busca clientes criados por um dado usuário."""
    return Cliente.objects.filter(criado_por=user)


class CepService:
    """Stub de serviço de busca de CEP."""

    @staticmethod
    def buscar(cep):
        # Retorna dados simulados para o CEP
        return {
            "cep": cep,
            "logradouro": "Rua Exemplo",
            "bairro": "Centro",
            "cidade": "Cidade Exemplo",
            "uf": "SP",
        }


class ClienteExportService:
    """Stub para exportação de clientes."""

    @staticmethod
    def exportar(clientes):
        return b""  # Retorna bytes vazios

    @staticmethod
    def exportar_excel(queryset: List[Cliente]) -> Workbook:
        """
        Exporta uma lista de clientes para um arquivo Excel.

        Args:
            queryset: QuerySet de clientes a serem exportados

        Returns:
            Workbook: Arquivo Excel com os dados dos clientes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Definir cabeçalhos
        headers = [
            "ID",
            "Nome",
            "CPF/CNPJ",
            "Email",
            "Telefone",
            "Tipo Cliente",
            "Status",
            "Data Cadastro",
            "Observação",
            "Endereço",
            "Número",
            "Complemento",
            "Bairro",
            "Cidade",
            "UF",
            "CEP",
        ]

        # Aplicar estilos aos cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Adicionar dados
        for row, cliente in enumerate(queryset, 2):
            data = [
                cliente.id,
                cliente.nome,
                getattr(cliente, "cpf_cnpj_formatado", cliente.cpf_cnpj),
                cliente.email,
                getattr(cliente, "telefone_formatado", cliente.telefone),
                cliente.tipo_cliente.nome if cliente.tipo_cliente else "",
                "Ativo" if cliente.ativo else "Inativo",
                cliente.data_cadastro.strftime("%d/%m/%Y %H:%M"),
                cliente.observacao,
                cliente.endereco,
                cliente.numero,
                cliente.complemento,
                cliente.bairro,
                cliente.cidade,
                cliente.uf,
                cliente.cep,
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(wrap_text=True)

        # Ajustar largura das colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Adicionar filtros
        ws.auto_filter.ref = ws.dimensions

        # Adicionar informações do relatório
        ws.sheet_properties.tabColor = "366092"

        return wb

    @staticmethod
    def get_filename() -> str:
        """Retorna o nome do arquivo para exportação."""
        return f'clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
